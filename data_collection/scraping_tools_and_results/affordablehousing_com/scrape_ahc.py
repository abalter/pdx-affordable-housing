"""
scrape_ahc.py  (AffordableHousing.com scraper)
===============================================
Uses the discovered JSON API to fetch all income-restricted Portland listings,
then fetches each detail page for rich data (fees, utilities, income tables, etc.)

SETUP (one-time):
  1. Open https://www.affordablehousing.com/portland-or/income-restricted-section8-owners/
     in your browser while logged out (or in a private window)
  2. Open DevTools → Network tab
  3. Reload the page
  4. Find the POST to /v4/AjaxHandler?message=SearchListings
  5. Right-click → Copy as cURL
  6. From that cURL, copy the full Cookie: header value into COOKIE below
  7. Copy the __RequestVerificationToken value into CSRF_TOKEN below

The cookies expire with your browser session, so you'll need to refresh them
if you run this more than a day later. Everything else stays the same.

Output:
    ahc_listings.json     - raw listing data (both pages, with GPS coords)
    ahc_details.json      - extracted detail data per property
    ahc_review.xlsx       - spreadsheet for review/merging

Usage:
    pip install requests beautifulsoup4 openpyxl
    python scrape_ahc.py
    python scrape_ahc.py --listings-only   # skip detail page fetching
    python scrape_ahc.py --details-only    # skip API, re-parse saved listings
    python scrape_ahc.py --excel-only      # just rebuild Excel from saved JSON
"""

import argparse
import json
import re
import time
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── !! UPDATE THESE BEFORE RUNNING !! ─────────────────────────────────────────
# Paste the full Cookie header value from the cURL command here:
COOKIE = (
    "ApplicationGatewayAffinity=9f5b645b4b20ee4389dfddcb0f7a7384; "
    "_vwo_uuid_v2=D1387FF6D1F0720B492EA8226023B77C5|caa3eb1587c4d4f708fc9a1b7277abaa; "
    "ASP.NET_SessionId=rmllbj4f11ibe53vcse4hj5b; "
    "SEOsearchResultParameters=%7B%22landlordUserId%22%3A0%2C%22state%22%3A%22or%22%2C%22county%22%3A%22%22%2C%22city%22%3A%22Portland%22%2C%22zip%22%3A%22%22%2C%22brand%22%3A%22%22%2C%22minPrice%22%3A0%2C%22maxPrice%22%3A0%2C%22hasSection8Voucher%22%3Afalse%2C%22yearlyIncome%22%3A3500%2C%22familySize%22%3A1%2C%22voucherSize%22%3A-1%2C%22isExcludeExceedsEligibility%22%3Afalse%2C%22bedCounts%22%3A%22%22%2C%22requiredMoreBedCounts%22%3A0%2C%22bathCount%22%3A0%2C%22sortExpression%22%3A%22LastUpdate%20Desc%22%2C%22itemsPerPage%22%3A32%2C%22page%22%3A1%2C%22returnOnlyCount%22%3Afalse%2C%22minLivingArea%22%3A0%2C%22maxLivingArea%22%3A0%2C%22propertyTypeCategories%22%3A%22%22%2C%22inUnitWasherAndDryer%22%3Afalse%2C%22onsiteLaundryFacilities%22%3Afalse%2C%22balconyPatio%22%3Afalse%2C%22parking%22%3Afalse%2C%22fitnessCenter%22%3Afalse%2C%22communitySwimmingPool%22%3Afalse%2C%22airConditioning%22%3Afalse%2C%22dishwasher%22%3Afalse%2C%22utilitiesIncluded%22%3Afalse%2C%22elevator%22%3Afalse%2C%22ageRestricted%22%3Afalse%2C%22noAgeRestrictions%22%3Afalse%2C%22petFriendly%22%3Afalse%2C%22leaseIncentives%22%3Afalse%2C%22picturesOnly%22%3Afalse%2C%22showSection8Badge%22%3Atrue%2C%22showTrustedOwnerBadge%22%3Afalse%2C%22incomeRestricted%22%3Atrue%2C%22physical%22%3Afalse%2C%22visualHearing%22%3Afalse%2C%22keywordSearch%22%3A%22%22%2C%22returnIdsOnly%22%3Afalse%2C%22isNewLocation%22%3A1%2C%22isIncludeMapListing%22%3Atrue%2C%22maxLatitude%22%3A%2245.81868004847237%22%2C%22maxLongitude%22%3A%22-122.2306207310562%22%2C%22minLatitude%22%3A%2245.2083534215771%22%2C%22minLongitude%22%3A%22-123.13877527350488%22%2C%22maxMarkerToShow%22%3A500%2C%22schoolId%22%3A0%2C%22NCESSchoolID%22%3A0%2C%22education%22%3A%22%22%2C%22pos%22%3A%2245.81868004847237%2C-122.2306207310562%2C45.2083534215771%2C-123.13877527350488%22%2C%22zoom%22%3A%228.826584719861712%22%2C%22center%22%3A%22-122.68469800228036%2C45.51434411601204%22%2C%22isNearOrMeetIncomeResultsOnly%22%3Afalse%2C%22enterpriseId%22%3A%22%22%7D; "
    "recent-match-data=%5B%7B%22stateCode%22%3A%22OR%22%2C%22stateName%22%3A%22Oregon%22%2C%22areaType%22%3A%22city%22%2C%22areaTitle%22%3A%22city%22%2C%22baseAreaName%22%3A%22Portland%22%2C%22areaName%22%3A%22Portland%2C%20OR%22%2C%22isExactMatch%22%3Afalse%2C%22matchCount%22%3A0%2C%22communityId%22%3A0%2C%22isCommunity%22%3Afalse%2C%22isLoggedUserCookie%22%3Afalse%7D%5D; "
    "cw_conversation=eyJhbGciOiJIUzI1NiJ9.eyJzb3VyY2VfaWQiOiI4Mjg3OGZiNy0yNmJiLTQxYWItYWIyYy1kMzNjZmE5NTgwZWYiLCJpbmJveF9pZCI6ODQyMTUsImV4cCI6MTc5MjA4NjM3MCwiaWF0IjoxNzc2NTM0MzcwfQ.Y4rBx2JLw_T7jNCh53kyo8LZRkEe7jH00Ec3bGuriVU; "
    "userlocation=%7B%22City%22%3A%22Milwaukie%22%2C%22County%22%3A%22Clackamas%20County%22%2C%22State%22%3A%22OR%22%2C%22Zip%22%3A%22%22%2C%22Latitude%22%3A%2245.4462%22%2C%22Longitude%22%3A%22-122.639%22%7D; "
    "PropertyGuidCookies=Guid=902108fc-635a-4a60-a6e3-1c1b45ae7813; "
    "__RequestVerificationToken=kiGX58WMZzJmCG71LFhQy_I6t7HN3TzHww4r3mGtfNQlasIjX1uraR4GwP9aI3RbOXQQMCMUByJRvsyZ7U5JOHXs0TM1; "
    "SEOTnResultPageVisited=true; "
    "affordable-filter=%7B%22filter%22%3A%7B%22isIncomeBasedMatch%22%3Afalse%2C%22isValidIncomeBasedMatch%22%3Afalse%2C%22isSection8Voucher%22%3Afalse%2C%22familySize%22%3A%221%22%2C%22voucherSize%22%3Anull%2C%22houseHoldIncome%22%3A%223%2C500%22%2C%22incomeFrequencyType%22%3A%22monthly%22%7D%7D; "
    "sourcesearch=true; "
    "userlastvalidsearchlocation=%7B%22City%22%3A%22Portland%22%2C%22State%22%3A%22or%22%2C%22Latitude%22%3A45.4667629%2C%22Longitude%22%3A-122.7109452%7D"
)

BASE_URL   = "https://www.affordablehousing.com"
API_URL    = BASE_URL + "/v4/AjaxHandler?message=SearchListings"
DETAIL_URL = BASE_URL + "{seo_url}"   # e.g. /portland-or/cascadian-terrace-491644/

LISTINGS_FILE = Path("ahc_listings.json")
DETAILS_FILE  = Path("ahc_details.json")
EXCEL_OUT     = Path("ahc_review.xlsx")

DELAY = 1.5   # seconds between requests

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:149.0) Gecko/20100101 Firefox/149.0",
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "Accept-Language": "en-US,en;q=0.9",
    "Content-Type": "application/json; charset=utf-8",
    "X-Requested-With": "XMLHttpRequest",
    "Referer": "https://www.affordablehousing.com/v4/pages/tnresult/tnresult.aspx",
    "Origin": BASE_URL,
    "Cookie": COOKIE,
}

# Base search parameters — we'll vary 'page' to paginate
SEARCH_PARAMS = {
    "landlordUserId": 0,
    "state": "or",
    "county": "",
    "city": "Portland",
    "zip": "",
    "brand": "",
    "minPrice": 0,
    "maxPrice": 0,
    "hasSection8Voucher": False,
    "yearlyIncome": 42000,  # ~$3500/mo — affects affordability calc only
    "familySize": 1,
    "voucherSize": -1,
    "isExcludeExceedsEligibility": False,
    "bedCounts": "",
    "requiredMoreBedCounts": 0,
    "bathCount": 0,
    "sortExpression": "LastUpdate Desc",
    "itemsPerPage": 32,
    "page": 1,
    "returnOnlyCount": False,
    "minLivingArea": 0,
    "maxLivingArea": 0,
    "propertyTypeCategories": "",
    "inUnitWasherAndDryer": False,
    "onsiteLaundryFacilities": False,
    "balconyPatio": False,
    "parking": False,
    "fitnessCenter": False,
    "communitySwimmingPool": False,
    "airConditioning": False,
    "dishwasher": False,
    "utilitiesIncluded": False,
    "elevator": False,
    "ageRestricted": False,
    "noAgeRestrictions": False,
    "petFriendly": False,
    "leaseIncentives": False,
    "picturesOnly": False,
    "showSection8Badge": True,
    "showTrustedOwnerBadge": False,
    "incomeRestricted": True,
    "physical": False,
    "visualHearing": False,
    "keywordSearch": "",
    "returnIdsOnly": False,
    "isNewLocation": 1,
    "isIncludeMapListing": True,
    # Bounding box covers Portland metro + a bit of buffer
    "maxLatitude": "45.818",
    "maxLongitude": "-122.230",
    "minLatitude": "45.208",
    "minLongitude": "-123.138",
    "maxMarkerToShow": 500,
    "schoolId": 0,
    "NCESSchoolID": 0,
    "education": "",
    "pos": "45.818,-122.230,45.208,-123.138",
    "zoom": "10",
    "center": "-122.684,45.514",
    "isNearOrMeetIncomeResultsOnly": False,
    "enterpriseId": "",
}


# ── Stage 1: Fetch all listing pages ─────────────────────────────────────────

def fetch_listings():
    all_listings = []
    page = 1

    print("── Stage 1: Fetching listing pages ───────────────────────────────")

    while True:
        params = {**SEARCH_PARAMS, "page": page}
        print(f"  Page {page}... ", end="", flush=True)

        try:
            r = requests.post(
                API_URL,
                json={"searchParameters": params},
                headers=HEADERS,
                timeout=15,
            )
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            print(f"FAILED: {e}")
            break

        listings = data.get("listings", [])
        total    = data.get("searchListingsParameters", {}).get("ItemCount", 0)
        per_page = data.get("searchListingsParameters", {}).get("ItemsPerPage", 32)

        if not listings:
            print("no results, done.")
            break

        print(f"{len(listings)} properties (total: {total})")
        all_listings.extend(listings)

        if len(all_listings) >= total:
            break

        page += 1
        time.sleep(DELAY)

    print(f"\n  ✓ {len(all_listings)} income-restricted properties found")

    # Normalize to clean dicts
    clean = []
    for p in all_listings:
        addr1 = p.get("AddressLine1","")
        addr2 = p.get("AddressLine2","")
        address = f"{addr1} {addr2}".strip() if addr2 else addr1

        clean.append({
            "community_id":  p.get("CommunityId"),
            "name":          p.get("CommunityName") or p.get("Title",""),
            "address":       address,
            "city":          p.get("City",""),
            "state":         p.get("State","OR"),
            "zip":           p.get("Zip",""),
            "lat":           p.get("Latitude"),
            "lon":           p.get("Longitude"),
            "min_rent":      p.get("MinAskingRent"),
            "max_rent":      p.get("MaxAskingRent"),
            "min_beds":      p.get("MinBedroomCount"),
            "max_beds":      p.get("MaxBedroomCount"),
            "min_sqft":      p.get("MinLivingArea"),
            "max_sqft":      p.get("MaxLivingArea"),
            "year_built":    p.get("YearBuilt"),
            "photo_url":     p.get("Photo",""),
            "availability":  p.get("AvailabilityText",""),
            "is_premium":    p.get("IsPremiumListing", False),
            "is_income_restricted": p.get("IsIncomeRestricted", True),
            "section8_badge": p.get("ShowSection8Badge", False),
            "seo_url":       p.get("SeoFriendlyRentalUrl",""),
            "description":   p.get("Description",""),
        })

    LISTINGS_FILE.write_text(json.dumps(clean, indent=2), encoding="utf-8")
    print(f"  ✓ Saved to {LISTINGS_FILE}\n")
    return clean


# ── Stage 2: Fetch and parse detail pages ─────────────────────────────────────

def parse_detail_page(html, seo_url):
    """
    Extract structured data using CSS selectors based on actual page structure.
    Sections use data-section="utilities|fees|details|amenities" attributes.
    """
    soup = BeautifulSoup(html, "html.parser")

    detail = {
        "seo_url":          seo_url,
        "phone":            None,
        "email":            None,
        "website_url":      None,
        "mgmt_company":     None,
        "units":            [],
        "fees": {
            "application":      None,
            "security_deposit": None,
            "pet_deposit":      None,
            "pet_fee":          None,
            "monthly_estimate": None,
            "movein_estimate":  None,
        },
        "utilities_tenant": [],
        "utilities_owner":  [],
        "amenities":        [],
        "pet_policy":       None,
        "age_restriction":  None,
        "renters_insurance": False,
        "raw_description":  None,
    }

    # Phone & Email
    phone_a = soup.find("a", href=re.compile(r"^tel:"))
    if phone_a:
        detail["phone"] = phone_a.get_text(strip=True)
    email_a = soup.find("a", href=re.compile(r"^mailto:"))
    if email_a:
        detail["email"] = email_a["href"].replace("mailto:", "").strip()

    # Management company ("Listed by Guardian Marketing")
    mgmt_m = re.search(r"Listed by\s+([^\n<]+)", html)
    if mgmt_m:
        detail["mgmt_company"] = mgmt_m.group(1).strip()

    # Property details section
    details_sec = soup.find("div", attrs={"data-section": "details"})
    if details_sec:
        for opt in details_sec.select("div.dtls--opt--cont"):
            opt_text = opt.get_text(" ", strip=True)
            spans = opt.select("span")
            smalls = opt.select("small")
            # Two structures:
            # Allowed: <span class="normal--fonts">Pets Allowed</span><small>(2 Max)</small>
            # Not allowed: <span>No</span><small>Pets Allowed</small>
            if smalls and spans:
                span_txt  = spans[0].get_text(strip=True)
                small_txt = smalls[0].get_text(strip=True)
                combined  = f"{span_txt} {small_txt}".lower()
                if "pet" in combined:
                    detail["pet_policy"] = opt_text
                elif "age" in combined:
                    detail["age_restriction"] = opt_text
        desc_el = details_sec.select_one("#communityDescription")
        if desc_el:
            desc_text = desc_el.get_text(" ", strip=True)
            detail["raw_description"] = desc_text[:500]
            m = re.search(r"go to:\s*(www\.[^\s,]+)", desc_text, re.I)
            if m:
                detail["website_url"] = "https://" + m.group(1).rstrip(".")

    # Units: div.CommunityModelRow
    for row in soup.select("div.CommunityModelRow"):
        cols = row.select("div.unit--col")
        if len(cols) < 4:
            continue
        # Remove <small> labels before extracting bed/bath counts
        for sm in cols[0].select("small"): sm.decompose()
        for sm in cols[1].select("small"): sm.decompose()
        beds  = cols[0].get_text(strip=True)
        baths = cols[1].get_text(strip=True)
        rent  = (row.select_one("div.rent--range--col") or BeautifulSoup("","html.parser")).get_text(strip=True)
        sqft  = (row.select_one("div.sqft--range--col") or BeautifulSoup("","html.parser")).get_text(strip=True)
        model = (row.select_one("span.model-text") or BeautifulSoup("","html.parser")).get_text(strip=True) or beds
        avail = (row.select_one("span.avl--text") or BeautifulSoup("","html.parser")).get_text(strip=True)
        if rent:
            detail["units"].append({"model": model, "beds": beds, "baths": baths,
                                    "rent": rent, "sqft": sqft, "availability": avail})

    # Fees section
    fees_sec = soup.find("div", attrs={"data-section": "fees"})
    if fees_sec:
        for li in fees_sec.select("li"):
            label_el = li.select_one("div.fees--col--sm p")
            price_el = li.select_one("div.fees--price strong")
            if not label_el:
                continue
            lbl   = label_el.get_text(strip=True).lower()
            price = price_el.get_text(strip=True) if price_el else None
            no_pets = bool(li.find(string=re.compile(r"no pets", re.I)))
            if "application fee" in lbl and price:
                detail["fees"]["application"] = price
            elif "security deposit" in lbl and price:
                detail["fees"]["security_deposit"] = price
            elif "pet deposit" in lbl:
                detail["fees"]["pet_deposit"] = None if no_pets else price
            elif "pet fee" in lbl and "monthly" not in lbl:
                detail["fees"]["pet_fee"] = None if no_pets else price
            elif "monthly estimate" in lbl and price:
                detail["fees"]["monthly_estimate"] = price
            elif "move-in estimate" in lbl and price:
                detail["fees"]["movein_estimate"] = price
        if fees_sec.find(string=re.compile(r"renters insurance", re.I)):
            detail["renters_insurance"] = True

    # Utilities section — note typo in class: "utilites--list"
    util_sec = soup.find("div", attrs={"data-section": "utilities"})
    if util_sec:
        util_cols = util_sec.select("div.utilities--col")
        if util_cols:
            detail["utilities_tenant"] = [
                li.get_text(strip=True) for li in util_cols[0].select("ul li")
                if li.get_text(strip=True)
            ]
        if len(util_cols) >= 2:
            detail["utilities_owner"] = [
                li.get_text(strip=True) for li in util_cols[1].select("ul li")
                if li.get_text(strip=True)
            ]

    # Amenities section — items WITHOUT line-through are available
    amen_sec = soup.find("div", attrs={"data-section": "amenities"})
    if amen_sec:
        detail["amenities"] = [
            li.get_text(strip=True)
            for li in amen_sec.select("ul.accessibility--list li")
            if "line-through" not in li.get("class", []) and li.get_text(strip=True)
        ]

    return detail



def fetch_details(listings):
    details = {}
    total = len(listings)

    print("── Stage 2: Fetching detail pages ────────────────────────────────")

    for i, p in enumerate(listings, 1):
        seo_url = p.get("seo_url","")
        if not seo_url:
            print(f"  [{i:3d}/{total}] {p.get('name','?')[:45]:<45} — no URL, skipping")
            continue

        url = BASE_URL + seo_url
        print(f"  [{i:3d}/{total}] {p.get('name','?')[:45]:<45}")

        try:
            r = requests.get(
                url,
                headers={**HEADERS,
                         "Accept": "text/html,application/xhtml+xml,*/*",
                         "Content-Type": ""},
                timeout=15
            )
            r.raise_for_status()
            detail = parse_detail_page(r.text, seo_url)
        except Exception as e:
            print(f"    ✗ {e}")
            detail = {"seo_url": seo_url, "error": str(e)}

        details[seo_url] = detail

        # Print summary
        units_str = ", ".join(f"{u['model']} {u['rent']}" for u in detail.get("units",[]))
        if units_str:
            print(f"    Rents: {units_str[:70]}")
        if detail.get("website_url"):
            print(f"    Site:  {detail['website_url']}")
        if detail.get("income_restrictions"):
            print(f"    Income restrictions: {len(detail['income_restrictions'])} tiers")

        time.sleep(DELAY)

    DETAILS_FILE.write_text(json.dumps(details, indent=2), encoding="utf-8")
    print(f"\n  ✓ Saved to {DETAILS_FILE}\n")
    return details


# ── Stage 3: Build Excel ──────────────────────────────────────────────────────

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, row, col, val, bg="1F4E79", fg="FFFFFF"):
    c = ws.cell(row, col, val)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(name="Arial", bold=True, color=fg, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _border()

def dat(ws, row, col, val, bg=None, wrap=False):
    c = ws.cell(row, col, str(val) if val is not None else "")
    c.font = Font(name="Arial", size=9)
    c.alignment = Alignment(vertical="top", wrap_text=wrap)
    c.border = _border()
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)

AVAIL_BG = {"Available Now": "D6F5E3", "Waiting List": "FEF9C3"}

def build_excel(listings, details):
    wb  = Workbook()
    ws  = wb.active
    ws.title = "AHC_Listings"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32

    cols = [
        "Name", "Address", "City", "ZIP", "Lat", "Lon",
        "Availability", "Min Rent", "Max Rent",
        "Min Beds", "Max Beds", "Min Sqft", "Max Sqft",
        "Year Built",
        "Phone", "Email", "Website", "Mgmt Company",
        "App Fee", "Security Dep", "Pet Fee",
        "Utilities (Tenant)", "Utilities (Owner)",
        "Income Restrictions",
        "Amenities", "Pet Policy", "Age Restriction",
        "Units Detail",
        "Description",
        "AHC URL", "Photo URL",
    ]
    for c, h in enumerate(cols, 1):
        hdr(ws, 1, c, h)

    for r, p in enumerate(listings, 2):
        seo  = p.get("seo_url","")
        det  = details.get(seo, {})
        fees = det.get("fees",{}) or {}
        avail = p.get("availability","")
        bg = AVAIL_BG.get(avail)

        units_str = " | ".join(
            f"{u['model']} {u['rent']} {u.get('sqft','')}sqft"
            for u in det.get("units",[])
        )
        income_str = " | ".join(
            f"Family {x['family_size']}: ${x['income_min']:,}-${x['income_max']:,} ({x['num_units']} units)"
            for x in det.get("income_restrictions",[])
        )

        vals = [
            p.get("name",""),
            p.get("address",""),
            p.get("city",""),
            p.get("zip",""),
            p.get("lat",""),
            p.get("lon",""),
            avail,
            f"${p.get('min_rent','')}" if p.get("min_rent") else "",
            f"${p.get('max_rent','')}" if p.get("max_rent") else "",
            p.get("min_beds",""),
            p.get("max_beds",""),
            p.get("min_sqft",""),
            p.get("max_sqft",""),
            p.get("year_built",""),
            det.get("phone",""),
            det.get("email",""),
            det.get("website_url",""),
            det.get("mgmt_company",""),
            fees.get("application",""),
            fees.get("security_deposit",""),
            fees.get("pet_fee",""),
            ", ".join(det.get("utilities_tenant",[])),
            ", ".join(det.get("utilities_owner",[])),
            income_str,
            ", ".join(det.get("amenities",[])),
            det.get("pet_policy",""),
            det.get("age_restriction",""),
            units_str,
            (p.get("description") or "")[:300],
            BASE_URL + seo,
            p.get("photo_url",""),
        ]

        for c, v in enumerate(vals, 1):
            wrap = c in (23, 24, 25, 26, 28, 29)
            dat(ws, r, c, v, bg=bg if c==7 else None, wrap=wrap)

    widths = [28,26,10,7,9,10, 12,10,10,8,8,8,8,10,
              14,24,30,20, 9,11,9, 24,22,36, 28,22,18, 40,40, 50,50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    wb.save(EXCEL_OUT)
    print(f"  ✓ Excel saved: {EXCEL_OUT}")
    avail_now = sum(1 for p in listings if p.get("availability") == "Available Now")
    print(f"  Available Now: {avail_now}  |  Waiting List: {len(listings)-avail_now}")


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--listings-only", action="store_true",
                        help="Only fetch listing pages, skip detail pages")
    parser.add_argument("--details-only",  action="store_true",
                        help="Skip API fetch, use saved listings + fetch details")
    parser.add_argument("--excel-only",    action="store_true",
                        help="Rebuild Excel from saved JSON files only")
    args = parser.parse_args()

    print(f"\n{'═'*60}")
    print(f"  AffordableHousing.com Scraper")
    print(f"{'═'*60}\n")

    if args.excel_only:
        listings = json.loads(LISTINGS_FILE.read_text())
        details  = json.loads(DETAILS_FILE.read_text()) if DETAILS_FILE.exists() else {}
        build_excel(listings, details)
        return

    if args.details_only:
        listings = json.loads(LISTINGS_FILE.read_text())
    else:
        listings = fetch_listings()

    if not args.listings_only:
        details = fetch_details(listings)
    else:
        details = {}

    build_excel(listings, details)


if __name__ == "__main__":
    main()

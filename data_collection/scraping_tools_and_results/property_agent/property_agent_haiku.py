"""
property_agent.py
=================
AI agent that finds property websites and extracts housing data.
Uses Claude's web_search tool (works from server IPs) with Haiku for all
steps to minimize cost.

Cost target: ~$1.84 total for 245 properties (vs ~$22 original)
  - Haiku replaces Sonnet for ALL steps (~20x cheaper)
  - web_search tool kept (works from any IP, unlike DuckDuckGo)

Usage:
    pip install anthropic requests beautifulsoup4 openpyxl
    export ANTHROPIC_API_KEY=sk-ant-...
    python property_agent.py --limit 5     # test run
    python property_agent.py               # full run (resume-safe)
    python property_agent.py --resume      # continue interrupted run
    python property_agent.py --excel-only  # rebuild Excel without API calls
"""

import argparse
import json
import os
import re
import time
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import anthropic
except ImportError:
    print("Missing: pip install anthropic"); raise

# ── Config ────────────────────────────────────────────────────────────────────

PROPS_JS     = Path("data/properties.js")
RESULTS_FILE = Path("agent_results.json")
EXCEL_OUT    = Path("agent_review.xlsx")

MODEL        = "claude-haiku-4-5-20251001"   # Haiku for all steps — ~20x cheaper
MAX_TOKENS   = 1024
FETCH_DELAY  = 0.8

HEADERS = {
    "User-Agent": (
        "pdx-affordable-housing/1.0 "
        "(nonprofit housing resource; github.com/abalter/pdx-affordable-housing)"
    )
}

SKIP_DOMAINS = {
    "apartments.com", "zillow.com", "trulia.com", "rent.com",
    "apartmentlist.com", "hotpads.com", "rentcafe.com", "realtor.com",
    "yelp.com", "google.com", "facebook.com", "twitter.com",
    "affordablehousingonline.com", "affordablehousinghub.org",
    "hud.gov", "oregon.gov", "portland.gov", "nwpilotproject.org",
    "homeforward.org", "211info.org", "linkedin.com", "youtube.com",
}


# ── Load / save ───────────────────────────────────────────────────────────────

def load_properties():
    text     = PROPS_JS.read_text(encoding="utf-8")
    json_str = re.sub(r"^\s*const\s+PROPS\s*=\s*", "", text).rstrip().rstrip(";")
    return json.loads(json_str)

def load_results():
    if RESULTS_FILE.exists():
        return json.loads(RESULTS_FILE.read_text(encoding="utf-8"))
    return {}

def save_results(results):
    RESULTS_FILE.write_text(
        json.dumps(results, indent=2, ensure_ascii=False), encoding="utf-8"
    )


# ── Fetch a page ──────────────────────────────────────────────────────────────

def fetch_page(url, max_chars=12000):
    """Fetch URL, return cleaned text truncated for token budget."""
    try:
        r = requests.get(url, headers=HEADERS, timeout=15, allow_redirects=True)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        for tag in soup(["script","style","nav","footer","header","aside"]):
            tag.decompose()
        text = soup.get_text(" ", strip=True)
        text = re.sub(r"\s{3,}", "\n", text)
        return text[:max_chars]
    except Exception as e:
        return f"FETCH_ERROR: {e}"


# ── Claude: web search ────────────────────────────────────────────────────────

def web_search(client, query):
    """
    Search using Anthropic's web_search tool on Haiku.
    Works from server IPs. Returns formatted results string.
    """
    try:
        messages = [{
            "role": "user",
            "content": (
                f"Search for: {query}\n\n"
                "List the top results with their URLs and brief descriptions. "
                "Focus on the official property website if one exists."
            )
        }]

        response = client.messages.create(
            model=MODEL,
            max_tokens=1500,
            tools=[{"type": "web_search_20250305", "name": "web_search"}],
            messages=messages,
        )

        # Handle tool use loop
        while response.stop_reason == "tool_use":
            tool_uses = [b for b in response.content if b.type == "tool_use"]
            tool_results = [
                {"type": "tool_result", "tool_use_id": tu.id, "content": ""}
                for tu in tool_uses
            ]
            messages.append({"role": "assistant", "content": response.content})
            messages.append({"role": "user",      "content": tool_results})
            response = client.messages.create(
                model=MODEL,
                max_tokens=1500,
                tools=[{"type": "web_search_20250305", "name": "web_search"}],
                messages=messages,
            )

        text = " ".join(
            b.text for b in response.content if hasattr(b, "text") and b.text
        ).strip()
        return text or "No results"

    except Exception as e:
        return f"SEARCH_ERROR: {e}"


# ── Claude: pick best URL ─────────────────────────────────────────────────────

def find_website(client, prop, search_text):
    """Haiku picks the best property URL from search results."""
    system = """Find the official website for an affordable housing property.
Prefer dedicated property sites. Accept management company portfolio pages for THIS property.
REJECT: apartments.com, zillow, trulia, yelp, facebook, hud.gov, oregon.gov, nwpilotproject.org.
If nothing suitable: return null.
JSON only, no markdown: {"url": "https://..." or null, "confidence": "high"|"medium"|"low"|"none", "reasoning": "brief"}"""

    msg = f"""Property: {prop['name']}
Address: {prop['address']}, {prop['city']}, {prop['state']} {prop['zip']}
Phone: {prop.get('phone','')}  Management: {prop.get('mgmt','')}

Search results:
{search_text[:2000]}"""

    try:
        resp = client.messages.create(
            model=MODEL, max_tokens=256,
            system=system,
            messages=[{"role":"user","content":msg}]
        )
        raw = re.sub(r"```json|```","", resp.content[0].text).strip()
        return json.loads(raw)
    except Exception:
        return {"url": None, "confidence": "none", "reasoning": "parse error"}


# ── Claude: extract data ──────────────────────────────────────────────────────

def extract_data(client, prop, page_text, url):
    """Haiku extracts structured data from a property page."""
    system = """Extract affordable housing data from a property website.
Only extract info explicitly on the page. Use null for missing fields.
JSON only, no markdown:
{"rents":{"studio":null,"one_br":null,"two_br":null,"three_br":null},
 "waitlist":{"status":"open"|"closed"|"unknown","notes":null},
 "availability":null,"phone":null,"email":null,
 "amenities":[],"eligibility":null,"accepts_vouchers":null,
 "pet_policy":null,"parking":null,"notes":null}"""

    msg = f"""Property: {prop['name']}
Address: {prop['address']}, {prop['city']}
URL: {url}

Page content:
{page_text[:6000]}"""

    try:
        resp = client.messages.create(
            model=MODEL, max_tokens=MAX_TOKENS,
            system=system,
            messages=[{"role":"user","content":msg}]
        )
        raw = re.sub(r"```json|```","", resp.content[0].text).strip()
        return json.loads(raw)
    except Exception as e:
        return {"parse_error": str(e)[:100]}


# ── Process one property ──────────────────────────────────────────────────────

def process_property(client, prop):
    result = {
        "id":             prop["id"],
        "name":           prop["name"],
        "address":        prop["address"],
        "city":           prop["city"],
        "program":        prop["program"],
        "mgmt":           prop.get("mgmt",""),
        "search_query":   "",
        "website_url":    None,
        "url_confidence": "none",
        "url_reasoning":  "",
        "extracted":      None,
        "error":          None,
    }

    # ── Step 1: Search ────────────────────────────────────────────────────
    mgmt  = prop.get("mgmt","")
    query = f'"{prop["name"]}" Portland Oregon apartments'
    if mgmt and len(mgmt) < 35:
        query += f' "{mgmt}"'
    result["search_query"] = query

    print(f"    🔍 {query[:72]}")
    search_text = web_search(client, query)
    print(f"    📄 {len(search_text)} chars")

    if "SEARCH_ERROR" in search_text:
        result["error"] = search_text
        return result

    # ── Step 2: Pick URL ──────────────────────────────────────────────────
    url_result = find_website(client, prop, search_text)
    url        = url_result.get("url")
    result["url_confidence"] = url_result.get("confidence","none")
    result["url_reasoning"]  = url_result.get("reasoning","")
    print(f"    🌐 {url or 'none'} [{result['url_confidence']}]")

    if not url:
        return result

    # Reject aggregators that slipped through
    domain = re.sub(r"https?://(www\.)?","", url).split("/")[0]
    if any(s in domain for s in SKIP_DOMAINS):
        print(f"    ⚠ Rejected: {domain}")
        result["url_confidence"] = "none"
        return result

    result["website_url"] = url

    # ── Step 3: Fetch ─────────────────────────────────────────────────────
    time.sleep(FETCH_DELAY)
    page_text = fetch_page(url)
    if "FETCH_ERROR" in page_text:
        result["error"] = page_text
        return result

    # ── Step 4: Extract ───────────────────────────────────────────────────
    extracted = extract_data(client, prop, page_text, url)
    result["extracted"] = extracted

    rents   = extracted.get("rents") or {}
    wl      = (extracted.get("waitlist") or {}).get("status","?")
    found   = [f"{k}:{v}" for k,v in rents.items() if v]
    print(f"    ✓ waitlist:{wl}  rents:{found or 'none'}")

    return result


# ── Main loop ─────────────────────────────────────────────────────────────────

def run_agent(limit=None, resume=True):
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise ValueError("Set ANTHROPIC_API_KEY environment variable")

    client     = anthropic.Anthropic(api_key=api_key)
    properties = load_properties()
    results    = load_results() if resume else {}

    if limit:
        properties = properties[:limit]

    todo  = [p for p in properties if str(p["id"]) not in results]
    done  = len(properties) - len(todo)

    print(f"\n{'═'*60}")
    print(f"  Property Website Agent  (model: {MODEL})")
    print(f"{'═'*60}")
    print(f"  Properties:   {len(properties)}")
    print(f"  Already done: {done}")
    print(f"  To process:   {len(todo)}")
    print(f"  Est. cost:    ~${len(todo)*0.0075:.2f}")
    print(f"{'═'*60}\n")

    for i, prop in enumerate(todo, 1):
        print(f"\n[{i}/{len(todo)}] {prop['name']} — {prop['address']}")
        try:
            result = process_property(client, prop)
        except Exception as e:
            print(f"    ✗ {e}")
            result = {
                "id":prop["id"],"name":prop["name"],
                "address":prop["address"],"city":prop["city"],
                "program":prop["program"],"error":str(e),
            }

        results[str(prop["id"])] = result
        save_results(results)

    found    = sum(1 for r in results.values() if r.get("website_url"))
    high     = sum(1 for r in results.values() if r.get("url_confidence")=="high")
    w_rent   = sum(1 for r in results.values()
                   if any((r.get("extracted") or {}).get("rents",{}).values()
                          if (r.get("extracted") or {}).get("rents") else []))
    wl_known = sum(1 for r in results.values()
                   if ((r.get("extracted") or {}).get("waitlist") or {}).get("status")
                   in ("open","closed"))

    print(f"\n{'═'*60}")
    print(f"  ✓ Done  →  {RESULTS_FILE}")
    print(f"  Website found:   {found}/{len(results)}  (high conf: {high})")
    print(f"  Rent data:       {w_rent}")
    print(f"  Waitlist known:  {wl_known}")
    print(f"{'═'*60}\n")
    return results


# ── Excel ─────────────────────────────────────────────────────────────────────

def _b():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, row, col, val, bg="1F4E79", fg="FFFFFF"):
    c = ws.cell(row, col, val)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(name="Arial", bold=True, color=fg, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _b(); return c

def dat(ws, row, col, val, bg=None, wrap=False):
    c = ws.cell(row, col, str(val) if val else "")
    c.font = Font(name="Arial", size=9)
    c.alignment = Alignment(vertical="top", wrap_text=wrap)
    c.border = _b()
    if bg: c.fill = PatternFill("solid", fgColor=bg)
    return c

CONF_BG = {"high":"D6F5E3","medium":"FEF9C3","low":"FFE8D0","none":"F0F0F0"}
WAIT_BG = {"open":"D6F5E3","closed":"FFE0E0","unknown":"F0F0F0"}

def build_excel(results, properties):
    wb = Workbook()
    pmap = {str(p["id"]): p for p in properties}

    # ── Summary sheet ─────────────────────────────────────────────────────
    ws = wb.active; ws.title = "Summary"
    ws.freeze_panes = "A2"; ws.row_dimensions[1].height = 32
    cols = ["Name","Address","City","Program","Mgmt",
            "Found","Confidence","Website URL",
            "Waitlist","Waitlist Notes",
            "Studio","1BR","2BR","3BR",
            "Availability","Phone","Email",
            "Vouchers","Eligibility","Amenities","Notes","Reasoning"]
    for c, h in enumerate(cols, 1): hdr(ws, 1, c, h)

    for r, (pid, res) in enumerate(sorted(results.items(), key=lambda x:int(x[0])), 2):
        ext  = res.get("extracted") or {}
        rents= ext.get("rents") or {}
        wl   = ext.get("waitlist") or {}
        conf = res.get("url_confidence","none")
        wstat= wl.get("status","unknown")
        rbg  = CONF_BG.get(conf)
        vals = [
            res.get("name",""), res.get("address",""), res.get("city",""),
            res.get("program",""), res.get("mgmt",""),
            "Yes" if res.get("website_url") else "No",
            conf, res.get("website_url",""),
            wstat, wl.get("notes",""),
            rents.get("studio",""), rents.get("one_br",""),
            rents.get("two_br",""), rents.get("three_br",""),
            ext.get("availability",""), ext.get("phone",""), ext.get("email",""),
            "Yes" if ext.get("accepts_vouchers") else
            ("No" if ext.get("accepts_vouchers") is False else ""),
            ext.get("eligibility",""),
            ", ".join(ext.get("amenities") or []),
            ext.get("notes",""), res.get("url_reasoning",""),
        ]
        for c, v in enumerate(vals, 1):
            bg = WAIT_BG.get(wstat) if c == 9 else rbg
            dat(ws, r, c, v, bg=bg, wrap=(c in (8,10,19,20,21,22)))

    for i, w in enumerate([26,22,10,12,18,6,10,38,10,28,10,10,10,10,16,14,24,7,22,28,22,28],1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    # ── No Website sheet ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("No_Website")
    ws2.freeze_panes = "A2"
    c2 = ["Name","Address","City","Program","Phone","Mgmt","Search Query","Reasoning"]
    for c, h in enumerate(c2, 1): hdr(ws2, 1, c, h, bg="555555")
    r = 2
    for pid, res in sorted(results.items(), key=lambda x:int(x[0])):
        if res.get("website_url"): continue
        p = pmap.get(pid, {})
        for c, v in enumerate([
            res.get("name",""), res.get("address",""), res.get("city",""),
            res.get("program",""), p.get("phone",""), res.get("mgmt",""),
            res.get("search_query",""), res.get("url_reasoning",""),
        ], 1): dat(ws2, r, c, v, wrap=(c>=7))
        r += 1
    for i, w in enumerate([26,22,10,12,14,20,42,32],1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Errors sheet ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Errors")
    ws3.freeze_panes = "A2"
    for c, h in enumerate(["Name","Address","Error"],1): hdr(ws3,1,c,h,bg="8B0000")
    r = 2
    for pid, res in sorted(results.items(), key=lambda x:int(x[0])):
        if not res.get("error"): continue
        for c, v in enumerate([res.get("name",""),res.get("address",""),res.get("error","")],1):
            dat(ws3, r, c, v, wrap=(c==3))
        r += 1
    for i, w in enumerate([26,22,60],1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    wb.save(EXCEL_OUT)
    print(f"  ✓ Excel: {EXCEL_OUT}")


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser()
    p.add_argument("--resume",     action="store_true", default=True)
    p.add_argument("--no-resume",  action="store_true")
    p.add_argument("--excel-only", action="store_true")
    p.add_argument("--limit",      type=int, default=None)
    args = p.parse_args()

    if not PROPS_JS.exists():
        print(f"✗ {PROPS_JS} not found — run from repo root"); return

    properties = load_properties()

    if args.excel_only:
        results = load_results()
        print(f"Rebuilding Excel from {len(results)} results...")
        build_excel(results, properties); return

    results = run_agent(limit=args.limit, resume=not args.no_resume)
    build_excel(results, properties)

if __name__ == "__main__":
    main()

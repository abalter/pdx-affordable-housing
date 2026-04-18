"""
property_agent.py
=================
AI agent that finds the actual website for each affordable housing property,
then extracts ground-truth data (rents, availability, waitlist, amenities, etc.)

Architecture:
  For each property:
    1. Web search: "{name}" "{city}" apartments site
    2. Claude evaluates results → picks best candidate URL (or none)
    3. Fetch candidate page
    4. Claude extracts structured data from the HTML
    5. Save to progress file (resume-safe)

Output:
    agent_results.json    - one record per property, incrementally saved
    agent_review.xlsx     - formatted workbook for review

Usage:
    pip install requests beautifulsoup4 openpyxl anthropic
    python property_agent.py

    # Resume interrupted run:
    python property_agent.py --resume

    # Just rebuild Excel from saved results (no API calls):
    python property_agent.py --excel-only

    # Limit to first N properties (for testing):
    python property_agent.py --limit 10

Notes:
    - Requires ANTHROPIC_API_KEY environment variable
    - Uses web_search tool via Anthropic API
    - ~2-3 API calls per property, ~245 properties → budget ~$2-4 total
    - Saves progress after every property so you can resume safely
"""

import argparse
import json
import os
import re
import time
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import anthropic
except ImportError:
    print("Missing: pip install anthropic")
    raise

# ── Config ────────────────────────────────────────────────────────────────────

PROPS_JS      = Path("data/properties.js")
RESULTS_FILE  = Path("agent_results.json")
EXCEL_OUT     = Path("agent_review.xlsx")

MODEL         = "claude-sonnet-4-5"
MAX_TOKENS    = 1024
FETCH_TIMEOUT = 15
FETCH_DELAY   = 0.8   # between page fetches (not API calls)

# Domains to skip — aggregators, not the actual property site
SKIP_DOMAINS = {
    "apartments.com", "zillow.com", "trulia.com", "rent.com",
    "apartmentlist.com", "hotpads.com", "rentcafe.com", "realtor.com",
    "yelp.com", "google.com", "facebook.com", "twitter.com",
    "affordablehousingonline.com", "affordablehousinghub.org",
    "hud.gov", "oregon.gov", "portland.gov", "nwpilotproject.org",
    "homeforward.org", "211info.org", "linkedin.com", "youtube.com",
}

HEADERS = {
    "User-Agent": (
        "pdx-affordable-housing/1.0 "
        "(nonprofit housing resource; github.com/abalter/pdx-affordable-housing)"
    )
}


# ── Load properties ───────────────────────────────────────────────────────────

def load_properties():
    text     = PROPS_JS.read_text(encoding="utf-8")
    json_str = re.sub(r"^\s*const\s+PROPS\s*=\s*", "", text).rstrip().rstrip(";")
    return json.loads(json_str)


# ── Load / save progress ──────────────────────────────────────────────────────

def load_results():
    if RESULTS_FILE.exists():
        return json.loads(RESULTS_FILE.read_text(encoding="utf-8"))
    return {}   # keyed by property id (as string)

def save_results(results):
    RESULTS_FILE.write_text(
        json.dumps(results, indent=2, ensure_ascii=False),
        encoding="utf-8"
    )


# ── Fetch a URL ───────────────────────────────────────────────────────────────

def fetch_page(url, max_chars=12000):
    """Fetch URL, return cleaned text (truncated for token budget)."""
    try:
        r = requests.get(url, headers=HEADERS, timeout=FETCH_TIMEOUT)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        # Remove nav, footer, scripts, styles
        for tag in soup(["script","style","nav","footer","header","aside"]):
            tag.decompose()
        text = soup.get_text(" ", strip=True)
        text = re.sub(r"\s{3,}", "\n", text)
        return text[:max_chars]
    except Exception as e:
        return f"FETCH_ERROR: {e}"


# ── Claude helpers ────────────────────────────────────────────────────────────

def claude(client, system, user_msg):
    """Single Claude call, returns response text."""
    response = client.messages.create(
        model=MODEL,
        max_tokens=MAX_TOKENS,
        system=system,
        messages=[{"role": "user", "content": user_msg}]
    )
    return response.content[0].text.strip()


def find_website(client, prop, search_results_text):
    """
    Ask Claude to pick the best property website URL from search results.
    Returns dict: {url, confidence, reasoning}
    """
    system = """You are helping find the official website for an affordable housing property.
Given property details and web search results, identify the best URL for the property's OWN website.

Rules:
- Prefer dedicated property sites (e.g. uptowntowerapartments.com)
- Accept management company portfolio pages if they're for THIS specific property
- REJECT aggregators: apartments.com, zillow, trulia, rent.com, yelp, facebook, etc.
- REJECT government/nonprofit directories: hud.gov, oregon.gov, nwpilotproject.org, etc.
- If no suitable URL exists, say so

Respond with valid JSON only, no markdown:
{
  "url": "https://..." or null,
  "confidence": "high" | "medium" | "low" | "none",
  "reasoning": "brief explanation"
}"""

    user_msg = f"""Property: {prop['name']}
Address: {prop['address']}, {prop['city']}, {prop['state']} {prop['zip']}
Phone: {prop.get('phone','')}
Management: {prop.get('mgmt','')}

Search results:
{search_results_text}"""

    raw = claude(client, system, user_msg)
    try:
        # Strip any accidental markdown fences
        raw = re.sub(r"```json|```", "", raw).strip()
        return json.loads(raw)
    except Exception:
        return {"url": None, "confidence": "none", "reasoning": f"Parse error: {raw[:100]}"}


def extract_property_data(client, prop, page_text, url):
    """
    Ask Claude to extract structured data from a property website's text.
    Returns dict with extracted fields.
    """
    system = """You are extracting affordable housing data from a property website.
Extract ONLY information explicitly present on the page. Do not guess or infer.
Use null for fields not found.

Respond with valid JSON only, no markdown:
{
  "website_url": "the URL",
  "rents": {
    "studio": "e.g. $850/mo or null",
    "one_br": "e.g. $1,050/mo or null",
    "two_br": "e.g. $1,200/mo or null",
    "three_br": "e.g. $1,450/mo or null"
  },
  "waitlist": {
    "status": "open" | "closed" | "unknown",
    "notes": "exact text from page if any, else null"
  },
  "availability": "e.g. '2 units available' or 'Call for availability' or null",
  "phone": "phone number if found, else null",
  "email": "email if found, else null",
  "amenities": ["list", "of", "amenities"] or [],
  "eligibility": "any eligibility text found (age, income, disability) or null",
  "accepts_vouchers": true | false | null,
  "pet_policy": "pet policy text or null",
  "parking": "parking info or null",
  "laundry": "laundry info or null",
  "notes": "anything else useful that doesn't fit above fields or null"
}"""

    user_msg = f"""Property: {prop['name']}
Address: {prop['address']}, {prop['city']}
URL: {url}

Page content:
{page_text}"""

    raw = claude(client, system, user_msg)
    try:
        raw = re.sub(r"```json|```", "", raw).strip()
        return json.loads(raw)
    except Exception:
        return {"parse_error": raw[:200], "website_url": url}


# ── Web search via Anthropic tool use ─────────────────────────────────────────

def web_search(client, query):
    """
    Use Anthropic's web_search tool to search for a property.
    Handles multi-turn tool use: user → tool_use → tool_result → final text.
    Returns formatted string of results for Claude to evaluate.
    """
    try:
        messages = [{
            "role": "user",
            "content": (
                f"Search for: {query}\n\n"
                "List the top results with their URLs and brief descriptions. "
                "Focus on the official property website if one exists."
            )
        }]

        # First call — Claude may invoke web_search tool
        response = client.messages.create(
            model=MODEL,
            max_tokens=2048,
            tools=[{"type": "web_search_20250305", "name": "web_search"}],
            messages=messages,
        )

        # If Claude used a tool, we need to feed results back and get final answer
        # Loop handles cases where Claude calls the tool multiple times
        while response.stop_reason == "tool_use":
            # Collect all tool use blocks
            tool_uses = [b for b in response.content if b.type == "tool_use"]

            # Build tool results — for web_search, Anthropic handles the actual
            # search server-side; we just echo back that it was called
            tool_results = []
            for tu in tool_uses:
                tool_results.append({
                    "type": "tool_result",
                    "tool_use_id": tu.id,
                    # web_search is server-side; content comes back in next response
                    "content": ""
                })

            # Append assistant response and tool results to messages
            messages.append({"role": "assistant", "content": response.content})
            messages.append({"role": "user",      "content": tool_results})

            # Continue conversation
            response = client.messages.create(
                model=MODEL,
                max_tokens=2048,
                tools=[{"type": "web_search_20250305", "name": "web_search"}],
                messages=messages,
            )

        # Collect all text blocks from final response
        result_text = []
        for block in response.content:
            if hasattr(block, "text") and block.text:
                result_text.append(block.text)

        output = "\n".join(result_text).strip()
        return output if output else "No results found"

    except Exception as e:
        return f"SEARCH_ERROR: {e}"


# ── Process one property ──────────────────────────────────────────────────────

def process_property(client, prop):
    """
    Full pipeline for one property.
    Returns result dict.
    """
    result = {
        "id":       prop["id"],
        "name":     prop["name"],
        "address":  prop["address"],
        "city":     prop["city"],
        "program":  prop["program"],
        "mgmt":     prop.get("mgmt", ""),
        "search_query":   "",
        "website_url":    None,
        "url_confidence": "none",
        "url_reasoning":  "",
        "extracted":      None,
        "error":          None,
    }

    # ── Step 1: Search ──────────────────────────────────────────────────
    # Try name + city first; if management company known, that's a useful signal too
    mgmt = prop.get("mgmt","")
    query = f'"{prop["name"]}" Portland Oregon apartments'
    if mgmt and len(mgmt) < 30:
        query += f' "{mgmt}"'
    result["search_query"] = query

    print(f"    🔍 Searching: {query[:70]}")
    search_text = web_search(client, query)
    print(f"    📄 Search returned {len(search_text)} chars: {search_text[:120]!r}")

    if "SEARCH_ERROR" in search_text:
        result["error"] = search_text
        return result

    # ── Step 2: Pick best URL ────────────────────────────────────────────
    url_result = find_website(client, prop, search_text)
    result["website_url"]    = url_result.get("url")
    result["url_confidence"] = url_result.get("confidence", "none")
    result["url_reasoning"]  = url_result.get("reasoning", "")

    print(f"    🌐 URL: {result['website_url'] or 'none'} [{result['url_confidence']}]")

    if not result["website_url"]:
        return result

    # Skip known aggregators that slipped through
    domain = re.sub(r"https?://(www\.)?", "", result["website_url"]).split("/")[0]
    if any(skip in domain for skip in SKIP_DOMAINS):
        print(f"    ⚠ Skipping aggregator domain: {domain}")
        result["website_url"]    = None
        result["url_confidence"] = "none"
        result["url_reasoning"]  = f"Rejected aggregator: {domain}"
        return result

    # ── Step 3: Fetch page ───────────────────────────────────────────────
    time.sleep(FETCH_DELAY)
    page_text = fetch_page(result["website_url"])

    if "FETCH_ERROR" in page_text:
        result["error"] = page_text
        return result

    # ── Step 4: Extract data ─────────────────────────────────────────────
    extracted = extract_property_data(client, prop, page_text, result["website_url"])
    result["extracted"] = extracted

    # Quick summary for console
    rents = extracted.get("rents", {}) or {}
    rent_found = [f"{k}:{v}" for k,v in rents.items() if v]
    waitlist = (extracted.get("waitlist") or {}).get("status", "?")
    print(f"    ✓ Extracted — waitlist:{waitlist}  rents:{rent_found or 'none found'}")

    return result


# ── Main scraping loop ────────────────────────────────────────────────────────

def run_agent(limit=None, resume=True):
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise ValueError("Set ANTHROPIC_API_KEY environment variable")

    client     = anthropic.Anthropic(api_key=api_key)
    properties = load_properties()
    results    = load_results() if resume else {}

    if limit:
        properties = properties[:limit]

    already_done = len([p for p in properties if str(p["id"]) in results])
    todo         = [p for p in properties if str(p["id"]) not in results]

    print(f"\n{'═'*60}")
    print(f"  Property Website Agent")
    print(f"{'═'*60}")
    print(f"  Total properties:  {len(properties)}")
    print(f"  Already done:      {already_done}")
    print(f"  To process:        {len(todo)}")
    print(f"{'═'*60}\n")

    for i, prop in enumerate(todo, 1):
        print(f"\n[{i}/{len(todo)}] {prop['name']} — {prop['address']}, {prop['city']}")
        try:
            result = process_property(client, prop)
        except Exception as e:
            print(f"    ✗ Unexpected error: {e}")
            result = {
                "id": prop["id"], "name": prop["name"],
                "address": prop["address"], "city": prop["city"],
                "program": prop["program"], "error": str(e),
            }

        results[str(prop["id"])] = result
        save_results(results)   # save after every property

    print(f"\n{'═'*60}")
    print(f"  Done. Results saved to {RESULTS_FILE}")

    # Summary stats
    found     = sum(1 for r in results.values() if r.get("website_url"))
    extracted = sum(1 for r in results.values() if r.get("extracted"))
    high_conf = sum(1 for r in results.values() if r.get("url_confidence") == "high")
    print(f"  Website found:     {found}/{len(results)}")
    print(f"  High confidence:   {high_conf}")
    print(f"  Data extracted:    {extracted}")
    print(f"{'═'*60}\n")

    return results


# ── Build Excel ───────────────────────────────────────────────────────────────

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, row, col, val, bg="1F4E79", fg="FFFFFF"):
    c = ws.cell(row, col, val)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(name="Arial", bold=True, color=fg, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()
    return c

def dat(ws, row, col, val, bg=None, wrap=False):
    c = ws.cell(row, col, val or "")
    c.font = Font(name="Arial", size=9)
    c.alignment = Alignment(vertical="top", wrap_text=wrap)
    c.border = thin_border()
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    return c

CONF_COLORS = {
    "high":   "D6F5E3",
    "medium": "FEF9C3",
    "low":    "FFE8D0",
    "none":   "F0F0F0",
}
WAIT_COLORS = {
    "open":    "D6F5E3",
    "closed":  "FFE0E0",
    "unknown": "F0F0F0",
}

def build_excel(results, properties):
    wb  = Workbook()
    props_by_id = {str(p["id"]): p for p in properties}

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 1: Summary — one row per property
    # ══════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.freeze_panes = "A2"
    ws1.row_dimensions[1].height = 32

    cols = [
        "Name", "Address", "City", "Program", "Mgmt",
        "Website Found", "Confidence", "Website URL",
        "Waitlist", "Waitlist Notes",
        "Studio Rent", "1BR Rent", "2BR Rent", "3BR Rent",
        "Availability", "Phone", "Email",
        "Accepts Vouchers", "Eligibility",
        "Amenities", "Notes", "URL Reasoning",
    ]
    for c, h in enumerate(cols, 1):
        hdr(ws1, 1, c, h)

    for r, (prop_id, res) in enumerate(sorted(results.items(), key=lambda x: int(x[0])), 2):
        prop = props_by_id.get(prop_id, {})
        ext  = res.get("extracted") or {}
        rents = ext.get("rents") or {}
        waitlist = ext.get("waitlist") or {}
        conf  = res.get("url_confidence", "none")
        wstat = waitlist.get("status", "unknown")

        row_bg = CONF_COLORS.get(conf)

        vals = [
            res.get("name",""),
            res.get("address",""),
            res.get("city",""),
            res.get("program",""),
            res.get("mgmt",""),
            "Yes" if res.get("website_url") else "No",
            conf,
            res.get("website_url",""),
            wstat,
            waitlist.get("notes",""),
            rents.get("studio",""),
            rents.get("one_br",""),
            rents.get("two_br",""),
            rents.get("three_br",""),
            ext.get("availability",""),
            ext.get("phone",""),
            ext.get("email",""),
            "Yes" if ext.get("accepts_vouchers") else ("No" if ext.get("accepts_vouchers") is False else ""),
            ext.get("eligibility",""),
            ", ".join(ext.get("amenities") or []),
            ext.get("notes",""),
            res.get("url_reasoning",""),
        ]

        for c, v in enumerate(vals, 1):
            # Waitlist column gets its own color
            bg = WAIT_COLORS.get(wstat) if c == 9 else row_bg
            dat(ws1, r, c, str(v) if v else "", bg=bg,
                wrap=(c in (8, 10, 19, 20, 21, 22)))

    widths = [28,24,10,12,18,8,10,40,10,30,12,12,12,12,18,14,24,8,24,30,24,30]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 2: No Website Found — properties where agent found nothing
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("No_Website")
    ws2.freeze_panes = "A2"
    ws2.row_dimensions[1].height = 32

    cols2 = ["Name","Address","City","Program","Phone","Mgmt","Search Query","Reasoning"]
    for c, h in enumerate(cols2, 1):
        hdr(ws2, 1, c, h, bg="555555")

    r = 2
    for prop_id, res in sorted(results.items(), key=lambda x: int(x[0])):
        if res.get("website_url"):
            continue
        dat(ws2, r, 1, res.get("name",""))
        dat(ws2, r, 2, res.get("address",""))
        dat(ws2, r, 3, res.get("city",""))
        dat(ws2, r, 4, res.get("program",""))
        dat(ws2, r, 5, props_by_id.get(prop_id, {}).get("phone",""))
        dat(ws2, r, 6, res.get("mgmt",""))
        dat(ws2, r, 7, res.get("search_query",""), wrap=True)
        dat(ws2, r, 8, res.get("url_reasoning",""), wrap=True)
        r += 1

    for i, w in enumerate([28,24,10,12,14,20,45,35], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 3: Errors
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Errors")
    ws3.freeze_panes = "A2"
    ws3.row_dimensions[1].height = 32

    cols3 = ["Name","Address","Error"]
    for c, h in enumerate(cols3, 1):
        hdr(ws3, 1, c, h, bg="8B0000")

    r = 2
    for prop_id, res in sorted(results.items(), key=lambda x: int(x[0])):
        if not res.get("error"):
            continue
        dat(ws3, r, 1, res.get("name",""))
        dat(ws3, r, 2, res.get("address",""))
        dat(ws3, r, 3, res.get("error",""), wrap=True)
        ws3.row_dimensions[r].height = 30
        r += 1

    for i, w in enumerate([28,24,60], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    wb.save(EXCEL_OUT)

    # Console summary
    total     = len(results)
    found     = sum(1 for r in results.values() if r.get("website_url"))
    high      = sum(1 for r in results.values() if r.get("url_confidence") == "high")
    with_rent = sum(1 for r in results.values()
                    if any((r.get("extracted") or {}).get("rents", {}).values()
                           if (r.get("extracted") or {}).get("rents") else []))
    waitlist_known = sum(1 for r in results.values()
                         if ((r.get("extracted") or {}).get("waitlist") or {}).get("status")
                         in ("open","closed"))

    print(f"\n  Excel saved to {EXCEL_OUT}")
    print(f"  ┌─────────────────────────────────┐")
    print(f"  │ Properties processed: {total:>4}      │")
    print(f"  │ Website found:        {found:>4}      │")
    print(f"  │   High confidence:    {high:>4}      │")
    print(f"  │ Rent data extracted:  {with_rent:>4}      │")
    print(f"  │ Waitlist known:       {waitlist_known:>4}      │")
    print(f"  └─────────────────────────────────┘\n")


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Property website agent")
    parser.add_argument("--resume",     action="store_true", default=True,
                        help="Skip already-processed properties (default: true)")
    parser.add_argument("--no-resume",  action="store_true",
                        help="Start fresh, ignore saved results")
    parser.add_argument("--excel-only", action="store_true",
                        help="Skip agent, just rebuild Excel from agent_results.json")
    parser.add_argument("--limit",      type=int, default=None,
                        help="Process only first N properties (for testing)")
    args = parser.parse_args()

    if not PROPS_JS.exists():
        print(f"✗ {PROPS_JS} not found — run from repo root")
        return

    properties = load_properties()

    if args.excel_only:
        if not RESULTS_FILE.exists():
            print(f"✗ {RESULTS_FILE} not found — run without --excel-only first")
            return
        results = load_results()
        print(f"Building Excel from {len(results)} saved results...")
        build_excel(results, properties)
        return

    resume = not args.no_resume
    results = run_agent(limit=args.limit, resume=resume)
    build_excel(results, properties)


if __name__ == "__main__":
    main()

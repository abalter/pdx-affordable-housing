"""
scrape_aho_to_excel.py
======================
Scrapes affordablehousingonline.com for Portland, OR affordable housing
and writes results to an Excel workbook for manual review.

Does NOT touch data/properties.js.

Sheets produced:
  1. AHO_Raw       - everything scraped from AHO, one row per property
  2. Matched        - side-by-side comparison: our data vs AHO data, for
                      properties that matched by address
  3. AHO_Only       - properties on AHO not found in our dataset
  4. Ours_Only      - our properties not found on AHO

Usage:
    pip install requests beautifulsoup4 openpyxl
    python scrape_aho_to_excel.py

    # To skip re-scraping and just rebuild Excel from saved raw data:
    python scrape_aho_to_excel.py --from-cache

Output:
    aho_raw.json          - raw scraped data (cache for re-runs)
    aho_review.xlsx       - workbook for manual review
"""

import argparse
import json
import re
import time
import unicodedata
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────

BASE_URL = "https://affordablehousingonline.com"
LIST_URL = BASE_URL + "/housing-search/Oregon/Portland"
DELAY    = 1.2
HEADERS  = {
    "User-Agent": (
        "pdx-affordable-housing/1.0 "
        "(nonprofit housing resource; github.com/abalter/pdx-affordable-housing)"
    )
}

RAW_CACHE  = Path("aho_raw.json")
PROPS_JS   = Path("data/properties.js")
EXCEL_OUT  = Path("aho_review.xlsx")


# ── HTTP helper ───────────────────────────────────────────────────────────────

def get(url, retries=3):
    for attempt in range(retries):
        try:
            r = requests.get(url, headers=HEADERS, timeout=15)
            r.raise_for_status()
            return r
        except Exception as e:
            print(f"  ⚠ Attempt {attempt+1} failed: {e}")
            time.sleep(3)
    return None


# ── Address normalizer (for fuzzy matching) ───────────────────────────────────

def normalize_address(addr):
    if not addr:
        return ""
    addr = addr.lower()
    addr = unicodedata.normalize("NFKD", addr).encode("ascii", "ignore").decode()
    addr = re.sub(r"[^\w\s]", " ", addr)
    for pattern, repl in [
        (r"\bstreet\b", "st"),   (r"\bavenue\b",    "ave"),
        (r"\bboulevard\b", "blvd"), (r"\bdrive\b",  "dr"),
        (r"\bplace\b",  "pl"),   (r"\broad\b",      "rd"),
        (r"\blane\b",   "ln"),   (r"\bnortheast\b", "ne"),
        (r"\bnorthwest\b", "nw"), (r"\bsoutheast\b", "se"),
        (r"\bsouthwest\b", "sw"), (r"\bnorth\b",    "n"),
        (r"\bsouth\b",  "s"),    (r"\beast\b",      "e"),
        (r"\bwest\b",   "w"),
    ]:
        addr = re.sub(pattern, repl, addr)
    return re.sub(r"\s+", " ", addr).strip()


# ── Stage 1: List pages ───────────────────────────────────────────────────────

def scrape_list_pages():
    stubs = []
    seen_ids = set()
    page = 1

    print("── Stage 1: Crawling list pages ──────────────────────────────")
    while True:
        url = LIST_URL if page == 1 else f"{LIST_URL}?page={page}"
        print(f"  Page {page}: ", end="", flush=True)
        r = get(url)
        if not r:
            print("failed, stopping.")
            break

        soup = BeautifulSoup(r.text, "html.parser")
        new = []
        for a in soup.select("a[href*='/housing-search/Oregon/Portland/']"):
            href = a.get("href", "")
            m = re.search(r"/housing-search/Oregon/Portland/([^/]+)/(\d+)", href)
            if not m:
                continue
            slug, aho_id = m.group(1), m.group(2)
            if aho_id in seen_ids:
                continue
            seen_ids.add(aho_id)
            h3 = a.find("h3")
            name = h3.get_text(strip=True) if h3 else slug.replace("-", " ").title()
            new.append({
                "name":       name,
                "detail_url": BASE_URL + f"/housing-search/Oregon/Portland/{slug}/{aho_id}",
                "aho_id":     aho_id,
                "slug":       slug,
            })

        if not new:
            print("no new properties, done.")
            break

        stubs.extend(new)
        print(f"{len(new)} properties (running total: {len(stubs)})")
        page += 1
        time.sleep(DELAY)

    print(f"\n  ✓ {len(stubs)} stubs\n")
    return stubs


# ── Stage 2: Detail pages ─────────────────────────────────────────────────────

def parse_detail(soup, stub):
    rec = {
        "aho_id":          stub["aho_id"],
        "aho_url":         stub["detail_url"],
        "name":            stub["name"],
        "address":         "",
        "city":            "",
        "state":           "OR",
        "zip":             "",
        "phone":           "",
        "email":           "",
        "waitlist":        "unknown",
        "rent_range":      "",
        "beds":            "",
        "photo_url":       "",
        "amenities":       "",
        "units_detail":    "",   # formatted unit table as string
        "accepts_vouchers": False,
        "accessible":      False,
        "mgmt":            "",
    }

    page_text = soup.get_text(" ", strip=True)
    page_lower = page_text.lower()

    # ── Address from heading area ──
    # The address sits right under the h1 property name
    h1 = soup.find("h1")
    if h1:
        # Look at the next sibling text nodes / elements
        for sib in h1.next_siblings:
            text = sib.get_text(strip=True) if hasattr(sib, "get_text") else str(sib).strip()
            if re.search(r"\d+\s+\w+.*OR\s+\d{5}", text):
                rec["address"] = text
                m = re.match(r"^(.*?),\s*([\w\s]+),\s*OR\s+(\d{5})", text)
                if m:
                    rec["address"] = m.group(1).strip()
                    rec["city"]    = m.group(2).strip()
                    rec["zip"]     = m.group(3)
                break

    # Fallback: search all text nodes
    if not rec["address"]:
        for tag in soup.find_all(["p", "div", "span"]):
            t = tag.get_text(strip=True)
            m = re.match(r"^(\d+\s+\S.*?),\s*([\w\s]+),\s*OR\s+(\d{5})", t)
            if m:
                rec["address"] = m.group(1).strip()
                rec["city"]    = m.group(2).strip()
                rec["zip"]     = m.group(3)
                break

    # ── Phone / Email ──
    phone = soup.find("a", href=re.compile(r"^tel:"))
    if phone:
        rec["phone"] = phone.get_text(strip=True)
    email = soup.find("a", href=re.compile(r"^mailto:"))
    if email:
        rec["email"] = email["href"].replace("mailto:", "").strip()

    # ── Management company ──
    mgmt_m = re.search(r"managed by ([^\n\.]+)", page_lower)
    if mgmt_m:
        rec["mgmt"] = mgmt_m.group(1).strip().title()

    # ── Waitlist ──
    if re.search(r"waitlist\s+(now\s+)?open|waiting list\s+(now\s+)?open", page_lower):
        rec["waitlist"] = "open"
    elif re.search(r"waitlist\s+is\s+closed|waiting list\s+is\s+closed|waitlist\s+closed", page_lower):
        rec["waitlist"] = "closed"
    elif re.search(r"short wait|likely short", page_lower):
        rec["waitlist"] = "short"
    elif re.search(r"long wait", page_lower):
        rec["waitlist"] = "long"

    # ── Rent range ──
    rent_m = re.search(r"\$([\d,]+)\s*[-–]\s*([\d,]+)\s*per\s*month", page_text, re.I)
    if rent_m:
        rec["rent_range"] = f"${rent_m.group(1)} - ${rent_m.group(2)}/mo"
    else:
        rent_m2 = re.search(r"\$([\d,]+)\s*per\s*month", page_text, re.I)
        if rent_m2:
            rec["rent_range"] = f"${rent_m2.group(1)}/mo"

    # ── Beds ──
    bed_m = re.search(r"(Studio|SRO|\d+)\s*[-–]?\s*(\d+)?\s*Beds?", page_text, re.I)
    if bed_m:
        rec["beds"] = bed_m.group(0).strip()

    # ── Unit table ──
    unit_lines = []
    for table in soup.find_all("table"):
        ths = [th.get_text(strip=True).lower() for th in table.find_all("th")]
        if "rent" in ths:
            for row in table.find_all("tr")[1:]:
                cells = [td.get_text(strip=True) for td in row.find_all("td")]
                if cells:
                    unit_lines.append(" | ".join(cells))
    rec["units_detail"] = "\n".join(unit_lines)

    # ── Photo ──
    img = soup.find("img", src=re.compile(r"s3\.amazonaws\.com.*\.(jpg|jpeg|png)", re.I))
    if img:
        rec["photo_url"] = img["src"]

    # ── Amenities ──
    amenities_h = soup.find(
        lambda t: t.name in ("h3","h4","h5","strong","b") and
                  "amenities" in t.get_text(strip=True).lower()
    )
    if amenities_h:
        ul = amenities_h.find_next("ul")
        if ul:
            rec["amenities"] = ", ".join(li.get_text(strip=True) for li in ul.find_all("li"))

    # ── Flags ──
    rec["accepts_vouchers"] = bool(re.search(r"accepts (housing )?vouchers|section 8", page_lower))
    rec["accessible"]       = bool(re.search(r"\baccessible\b|\bada\b", page_lower))

    return rec


def scrape_details(stubs):
    records = []
    total = len(stubs)
    print("── Stage 2: Fetching detail pages ────────────────────────────")
    for i, stub in enumerate(stubs, 1):
        print(f"  [{i:3d}/{total}] {stub['name'][:55]}")
        r = get(stub["detail_url"])
        if not r:
            print("    ✗ fetch failed")
            records.append({**stub, "fetch_error": True})
        else:
            records.append(parse_detail(BeautifulSoup(r.text, "html.parser"), stub))
        time.sleep(DELAY)
    print(f"\n  ✓ {len(records)} records\n")
    return records


# ── Stage 3: Build Excel workbook ────────────────────────────────────────────

# Style helpers
def hdr_cell(ws, row, col, value, bg="1F4E79", fg="FFFFFF", bold=True):
    c = ws.cell(row, col, value)
    c.fill   = PatternFill("solid", fgColor=bg)
    c.font   = Font(name="Arial", bold=bold, color=fg, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c

def data_cell(ws, row, col, value, bg=None, wrap=False):
    c = ws.cell(row, col, value)
    c.font      = Font(name="Arial", size=9)
    c.alignment = Alignment(vertical="top", wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    return c

THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

def apply_border(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row+1):
        for c in range(min_col, max_col+1):
            ws.cell(r, c).border = THIN


def load_existing():
    text = PROPS_JS.read_text(encoding="utf-8")
    json_str = re.sub(r"^\s*const\s+PROPS\s*=\s*", "", text).rstrip().rstrip(";")
    return json.loads(json_str)


def build_excel(aho_records, existing):
    wb = Workbook()

    # ── Build match index ──────────────────────────────────────────────────
    existing_index = {}   # norm_addr → existing prop
    for p in existing:
        addr = f"{p.get('address','')} {p.get('city','')} {p.get('state','OR')} {p.get('zip','')}".strip()
        key  = normalize_address(addr)
        if key:
            existing_index[key] = p

    matched   = []   # (aho_rec, our_prop)
    aho_only  = []   # on AHO, not in ours
    ours_only = list(existing)  # start with all ours; remove matched ones

    aho_matched_names = set()
    for rec in aho_records:
        if rec.get("fetch_error"):
            continue
        addr = f"{rec.get('address','')} {rec.get('city','Portland')} OR {rec.get('zip','')}".strip()
        key  = normalize_address(addr)
        if key in existing_index:
            matched.append((rec, existing_index[key]))
            aho_matched_names.add(existing_index[key].get("name",""))
        else:
            aho_only.append(rec)

    ours_only = [p for p in existing if p.get("name","") not in aho_matched_names]

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 1: AHO Raw
    # ══════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "AHO_Raw"
    ws1.freeze_panes = "A2"
    ws1.row_dimensions[1].height = 30

    cols1 = [
        "Name", "Address", "City", "ZIP", "Phone", "Email",
        "Waitlist", "Rent Range", "Beds", "Units Detail",
        "Accepts Vouchers", "Accessible", "Mgmt", "Amenities",
        "Photo URL", "AHO URL", "AHO ID",
    ]
    for c, h in enumerate(cols1, 1):
        hdr_cell(ws1, 1, c, h)

    for r, rec in enumerate(aho_records, 2):
        if rec.get("fetch_error"):
            data_cell(ws1, r, 1, rec.get("name",""), bg="FFE0E0")
            data_cell(ws1, r, 2, "FETCH ERROR")
            continue
        vals = [
            rec.get("name",""),
            rec.get("address",""),
            rec.get("city",""),
            rec.get("zip",""),
            rec.get("phone",""),
            rec.get("email",""),
            rec.get("waitlist",""),
            rec.get("rent_range",""),
            rec.get("beds",""),
            rec.get("units_detail",""),
            "Yes" if rec.get("accepts_vouchers") else "",
            "Yes" if rec.get("accessible") else "",
            rec.get("mgmt",""),
            rec.get("amenities",""),
            rec.get("photo_url",""),
            rec.get("aho_url",""),
            rec.get("aho_id",""),
        ]
        # Color-code waitlist
        wl = rec.get("waitlist","unknown")
        row_bg = None
        if wl == "open":   row_bg = "D6F5E3"
        elif wl == "closed": row_bg = "FFE0E0"
        elif wl == "short":  row_bg = "FEF9C3"

        for c, v in enumerate(vals, 1):
            cell = data_cell(ws1, r, c, v, bg=row_bg,
                             wrap=(c in (10, 14)))  # wrap units_detail and amenities

    col_widths1 = [28,28,12,8,14,24,10,14,12,30,8,8,20,30,40,50,10]
    for i, w in enumerate(col_widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    apply_border(ws1, 1, len(aho_records)+1, 1, len(cols1))
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(cols1))}1"

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 2: Matched (side-by-side diff)
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Matched")
    ws2.freeze_panes = "A2"
    ws2.row_dimensions[1].height = 30

    # Two-level header: group label + field label
    groups = [
        ("",           ["Name"]),
        ("Our Data",   ["Address","Phone","Rent Studio","Rent 1BR","Rent 2BR","Program","Mgmt","Age Elig"]),
        ("AHO Data",   ["Address","Phone","Rent Range","Beds","Waitlist","Email","Mgmt","Accessible","Vouchers","AHO URL"]),
        ("",           ["Match Quality"]),
    ]
    col = 1
    for group_label, fields in groups:
        span = len(fields)
        bg = "1F4E79" if not group_label else ("2E6B9E" if group_label == "Our Data" else "1a7a4a")
        for f in fields:
            hdr_cell(ws2, 1, col, (group_label + "\n" + f) if group_label else f,
                     bg=bg if group_label else "1F4E79")
            col += 1

    for r, (aho, our) in enumerate(matched, 2):
        our_addr = f"{our.get('address','')} {our.get('city','')}".strip()
        aho_addr = f"{aho.get('address','')} {aho.get('city','')}".strip()

        # Simple match quality heuristic
        def addr_sim(a, b):
            a_norm, b_norm = normalize_address(a), normalize_address(b)
            a_words, b_words = set(a_norm.split()), set(b_norm.split())
            if not a_words: return 0
            return len(a_words & b_words) / len(a_words | b_words)

        sim = addr_sim(our_addr, aho_addr)
        match_q = "✓ Strong" if sim > 0.7 else ("~ Partial" if sim > 0.4 else "? Weak")
        match_bg = "D6F5E3" if sim > 0.7 else ("FEF9C3" if sim > 0.4 else "FFE0E0")

        vals = [
            our.get("name",""),
            our_addr,
            our.get("phone",""),
            our.get("rentStudio",""),
            our.get("rent1br",""),
            our.get("rent2br",""),
            our.get("program",""),
            our.get("mgmt",""),
            our.get("ageElig",""),
            aho_addr,
            aho.get("phone",""),
            aho.get("rent_range",""),
            aho.get("beds",""),
            aho.get("waitlist",""),
            aho.get("email",""),
            aho.get("mgmt",""),
            "Yes" if aho.get("accessible") else "",
            "Yes" if aho.get("accepts_vouchers") else "",
            aho.get("aho_url",""),
            match_q,
        ]
        for c, v in enumerate(vals, 1):
            bg = match_bg if c == len(vals) else None
            data_cell(ws2, r, c, v, bg=bg)

    col_widths2 = [28,28,14,10,10,10,12,20,14, 28,14,14,10,10,24,20,8,8,40, 12]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    apply_border(ws2, 1, len(matched)+1, 1, len(col_widths2))
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(col_widths2))}1"

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 3: AHO Only (not in our data)
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("AHO_Only")
    ws3.freeze_panes = "A2"
    ws3.row_dimensions[1].height = 30

    cols3 = ["Name","Address","City","ZIP","Phone","Email","Waitlist",
             "Rent Range","Beds","Accepts Vouchers","Accessible","Amenities","AHO URL"]
    for c, h in enumerate(cols3, 1):
        hdr_cell(ws3, 1, c, h, bg="a05c10")

    for r, rec in enumerate(aho_only, 2):
        vals = [
            rec.get("name",""), rec.get("address",""), rec.get("city",""),
            rec.get("zip",""),  rec.get("phone",""),   rec.get("email",""),
            rec.get("waitlist",""), rec.get("rent_range",""), rec.get("beds",""),
            "Yes" if rec.get("accepts_vouchers") else "",
            "Yes" if rec.get("accessible") else "",
            rec.get("amenities",""), rec.get("aho_url",""),
        ]
        for c, v in enumerate(vals, 1):
            data_cell(ws3, r, c, v)

    col_widths3 = [30,28,12,8,14,26,10,14,12,8,8,35,50]
    for i, w in enumerate(col_widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    apply_border(ws3, 1, len(aho_only)+1, 1, len(cols3))
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(cols3))}1"

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 4: Ours Only (not found on AHO)
    # ══════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Ours_Only")
    ws4.freeze_panes = "A2"
    ws4.row_dimensions[1].height = 30

    cols4 = ["Name","Address","City","ZIP","Phone","Program","Age Eligibility","Region","Mgmt"]
    for c, h in enumerate(cols4, 1):
        hdr_cell(ws4, 1, c, h, bg="555555")

    for r, p in enumerate(ours_only, 2):
        vals = [
            p.get("name",""),    p.get("address",""), p.get("city",""),
            p.get("zip",""),     p.get("phone",""),   p.get("program",""),
            p.get("ageElig",""), p.get("region",""),  p.get("mgmt",""),
        ]
        for c, v in enumerate(vals, 1):
            data_cell(ws4, r, c, v)

    col_widths4 = [30,28,12,8,14,14,18,18,22]
    for i, w in enumerate(col_widths4, 1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    apply_border(ws4, 1, len(ours_only)+1, 1, len(cols4))
    ws4.auto_filter.ref = f"A1:{get_column_letter(len(cols4))}1"

    # ── Summary numbers on Sheet 1 ──
    ws1["A1"].comment = None  # just in case
    # Add a small summary block above the data (insert rows)
    # Actually keep it simple - just print to console

    wb.save(EXCEL_OUT)

    return {
        "aho_total":   len(aho_records),
        "matched":     len(matched),
        "aho_only":    len(aho_only),
        "ours_only":   len(ours_only),
    }


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--from-cache", action="store_true",
                        help="Skip scraping, rebuild Excel from aho_raw.json")
    args = parser.parse_args()

    print("\n╔══════════════════════════════════════════════════════════╗")
    print("║  AHO → Excel  (review before merging)                    ║")
    print("╚══════════════════════════════════════════════════════════╝\n")

    if args.from_cache:
        print(f"Loading cached data from {RAW_CACHE}...")
        records = json.loads(RAW_CACHE.read_text(encoding="utf-8"))
        print(f"  {len(records)} records loaded\n")
    else:
        stubs   = scrape_list_pages()
        records = scrape_details(stubs)
        RAW_CACHE.write_text(json.dumps(records, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"✓ Raw data cached to {RAW_CACHE}\n")

    if not PROPS_JS.exists():
        print(f"✗ {PROPS_JS} not found — run from repo root")
        return

    print("── Stage 3: Building Excel workbook ──────────────────────────")
    existing = load_existing()
    print(f"  Loaded {len(existing)} existing properties")

    stats = build_excel(records, existing)

    print(f"\n  ✓ Saved to {EXCEL_OUT}")
    print(f"\n  Summary:")
    print(f"    AHO properties scraped:       {stats['aho_total']}")
    print(f"    Matched to our data:          {stats['matched']}")
    print(f"    On AHO, not in ours:          {stats['aho_only']}  ← Sheet: AHO_Only")
    print(f"    In ours, not on AHO:          {stats['ours_only']}  ← Sheet: Ours_Only")
    print(f"\n  Sheets: AHO_Raw · Matched · AHO_Only · Ours_Only")
    print(f"\n  Next step: review {EXCEL_OUT}, then run merge_aho.py when ready.\n")


if __name__ == "__main__":
    main()
